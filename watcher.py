"""
watcher.py — Reporte de Cosecha AR
Lee Reporte de Cosecha ARG - Database.xlsx y pushea harvest_data.json al repo via GitHub API.
"""

import json
import time
import base64
import hashlib
import urllib.request
import urllib.error
from pathlib import Path
from openpyxl import load_workbook
import os

# ── CONFIG ──────────────────────────────────────────────────────────────────
EXCEL_PATH   = Path(__file__).parent / "Reporte de Cosecha ARG - Database.xlsx"
GITHUB_TOKEN = "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"   # mismo token
GITHUB_OWNER = "gcuki11-bit"
GITHUB_REPO  = "harvest-dashboard-ar"       # ← repo nuevo para AR
GITHUB_FILE  = "public/harvest_data.json"
POLL_SECONDS = 30

# ── CLIENTES conocidos en AR (Bin Code = destino externo) ───────────────────
KNOWN_CLIENTS = {
    "MOLINOS", "BUNGE", "CARGILL", "LDC", "FYO", "BAYER", "BAYA",
    "GENTOS", "LOS NOGALES", "MONTIELERO", "PGG W", "ZINMA (FORRATEC)",
}

def read_excel():
    wb = load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)

    # ── HARVEST MATRIX ──────────────────────────────────────────────────────
    harvest_rows = []
    ws = wb["Harvest matrix"]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2: continue
        if not row[1]: continue
        harvest_rows.append({
            "farm":       row[1],
            "crop":       row[2],
            "irrigated":  row[3],
            "plot":       str(row[4]) if row[4] else "",
            "ha_total":   row[5],
            "ha_done":    row[6],
            "ha_rem":     row[7],
            "prod_t":     row[8],
            "yield_real": row[9],
            "yield_est":  row[10],
            "prod_est":   row[11],
            "diff_pct":   row[12],
            "price_cur":  row[13],
            "amount_cur": row[14],
            "price_est":  row[15],
            "amount_est": row[16],
            "job_no":     row[17],
        })

    # ── CARTAS DE PORTE ──────────────────────────────────────────────────────
    movements = []
    ws2 = wb["Cartas de porte"]
    for i, row in enumerate(ws2.iter_rows(values_only=True)):
        if i < 2: continue
        if not row[1]: continue

        net_dest   = (row[8]  or 0) / 1000
        net_origin = (row[18] or 0) / 1000
        hum_kg     = row[9]  or 0
        hum_pct    = row[10] or 0
        cont_kg    = row[11] or 0
        cont_pct   = row[12] or 0

        has_dest  = bool(row[8])
        qty_tn    = net_dest if has_dest else net_origin
        descuento = round(net_origin - net_dest, 4) if has_dest else 0

        date_val = row[3]
        date_str = date_val.strftime("%Y-%m-%d") if date_val else ""

        movements.append({
            "ticket":        str(row[1]) if row[1] else "",
            "date":          date_str,
            "farm":          row[4] or "",
            "item":          row[6] or "",
            "bin_code":      row[24] or "",
            "qty_tn":        round(qty_tn, 4),
            "qty_origin_tn": round(net_origin, 4),
            "has_dest":      has_dest,
            "humidity_kg":   hum_kg,
            "humidity_pct":  hum_pct,
            "contam_kg":     cont_kg,
            "contam_pct":    cont_pct,
            "descuento_tn":  descuento,
        })

    return {"harvest": harvest_rows, "movements": movements}


def get_file_sha():
    url = f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/contents/{GITHUB_FILE}"
    req = urllib.request.Request(url, headers={
        "Authorization": f"token {GITHUB_TOKEN}",
        "Accept": "application/vnd.github.v3+json",
    })
    try:
        with urllib.request.urlopen(req) as r:
            return json.loads(r.read())["sha"]
    except urllib.error.HTTPError as e:
        if e.code == 404:
            return None
        raise


def push_to_github(data):
    content = json.dumps(data, ensure_ascii=False, indent=2)
    encoded = base64.b64encode(content.encode("utf-8")).decode("ascii")
    sha = get_file_sha()

    url     = f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/contents/{GITHUB_FILE}"
    payload = {
        "message": "watcher: update harvest_data.json",
        "content": encoded,
    }
    if sha:
        payload["sha"] = sha

    body = json.dumps(payload).encode("utf-8")
    req  = urllib.request.Request(url, data=body, method="PUT", headers={
        "Authorization":  f"token {GITHUB_TOKEN}",
        "Accept":         "application/vnd.github.v3+json",
        "Content-Type":   "application/json",
    })
    with urllib.request.urlopen(req) as r:
        return r.status


def main():
    print("🌾 Watcher AR iniciado — monitoreando Excel cada", POLL_SECONDS, "seg")
    last_hash = None

    while True:
        try:
            if not EXCEL_PATH.exists():
                print(f"  ⚠ No se encontró el Excel en: {EXCEL_PATH}")
                time.sleep(POLL_SECONDS)
                continue

            raw   = EXCEL_PATH.read_bytes()
            h     = hashlib.md5(raw).hexdigest()

            if h != last_hash:
                print(f"  📊 Cambio detectado — leyendo Excel...")
                data   = read_excel()
                status = push_to_github(data)
                print(f"  ✓ Push OK (HTTP {status}) — {len(data['harvest'])} filas harvest, "
                      f"{len(data['movements'])} movimientos")
                last_hash = h
            else:
                print("  · Sin cambios")

        except Exception as e:
            print(f"  ✗ Error: {e}")

        time.sleep(POLL_SECONDS)


if __name__ == "__main__":
    main()
